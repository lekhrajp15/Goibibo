import logging
from configparser import ConfigParser
import openpyxl




# def getlog():
#     logger = logging.getLogger(__name__)
#     filehandler= logging.FileHandler('logfile.log')
#     formatter = logging.Formatter("%(asctime)s: %(levelname)s: %(name)s:%(message)s")
#     filehandler.setFormatter(formatter)
#     logger.addHandler(filehandler)

def readconfigfile(path, environment, value):
    cp = ConfigParser()
    cp.read(path)
    return cp.get(environment, value)

def writeinconfile(path, environment, key, value):
    cp = ConfigParser()
    cp[environment] = {key: value}
    with open(path, 'w') as configfile:
        cp.write(configfile)

def getmaxrow(path, sheetname):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetname]
    return sh.max_row

def getmaxcolumn(path, sheetname):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetname]
    return sh.max_column

def getcelldata(path, sheetname, row, column):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetname]
    return sh.cell(row=row, column=column).value


def setcelldata(path, sheetname, row, column,data):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheetname]
    sh.cell(row=row, column=column).value = data
    wb.save(path)







